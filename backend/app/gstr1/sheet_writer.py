import shutil
from typing import List, Type

import pandas as pd
from openpyxl import load_workbook

from app.gstr1.sheet_builders.b2b import B2BBuilder
from app.gstr1.sheet_builders.b2cl import B2CLBuilder
from app.gstr1.sheet_builders.b2cs import B2CSBuilder
from app.gstr1.sheet_builders.cdnr import CDNRBuilder
from app.gstr1.sheet_builders.cdnur import CDNURBuilder
from app.gstr1.sheet_builders.exp import EXPBuilder
from app.gstr1.sheet_builders.hsnb2b import HSNB2BBuilder
from app.gstr1.sheet_builders.hsnb2c import HSNB2CBuilder  
from app.gstr1.sheet_builders.eco import ECOBuilder  
from app.gstr1.sheet_builders.ecob2b import ECOB2BBuilder 

class SheetWriter:
    """
    Writes output into a COPY of the GST template using openpyxl only.
    This preserves all dropdowns, data validations, styles, merged cells, etc.
    """

    HEADER_ROW_INDEX = 3      # Excel row 4
    DATA_START_ROW_INDEX = 4  # Excel row 5 (0-based for DataFrame, but openpyxl uses 1-based)

    def __init__(self, template_path: str, output_path: str):
        self.template_path = template_path
        self.output_path = output_path

        # Copy template before we start writing
        shutil.copyfile(self.template_path, self.output_path)

        self.builder_classes: List[Type] = [
            B2BBuilder,
            B2CLBuilder,
            B2CSBuilder,
            CDNRBuilder,
            CDNURBuilder,
            EXPBuilder,
            HSNB2BBuilder,  
            HSNB2CBuilder,
            # ECOBuilder,
            # ECOB2BBuilder,
        ]


    def _get_sheet_headers(self, ws) -> List[str]:
        """
        Reads header row from the template (usually row 4).
        """
        row = ws[self.HEADER_ROW_INDEX + 1]  # openpyxl = 1-based
        return [cell.value for cell in row]

    def _write_sheet_data(self, ws, df: pd.DataFrame, headers: List[str]):
        """
        Writes DataFrame values into the openpyxl worksheet below the header row.
        """
        if df.empty:
            return

        start_row = self.DATA_START_ROW_INDEX + 1  # Excel row index (5)

        for r_idx, row in df.iterrows():
            excel_row = start_row + r_idx
            for c_idx, header in enumerate(headers, start=1):
                ws.cell(row=excel_row, column=c_idx, value=row.get(header))

    def build_and_write_all(self, df: pd.DataFrame) -> None:
        """
        Build data for each sheet and write directly to the copied template.
        """

        wb = load_workbook(self.output_path, keep_links=False)

        for builder_cls in self.builder_classes:
            builder = builder_cls()
            sheet_name = builder.SHEET_NAME

            if sheet_name not in wb.sheetnames:
                print(f"Skipping missing sheet: {sheet_name}")
                continue

            ws = wb[sheet_name]
            headers = self._get_sheet_headers(ws)

            sheet_df = builder.build(df, headers)
            if sheet_df is None:
                continue

            # Ensure correct column order
            final_df = pd.DataFrame(columns=headers)
            for col in sheet_df.columns:
                if col in final_df.columns:
                    final_df[col] = sheet_df[col]

            self._write_sheet_data(ws, final_df, headers)

            # ✅ GENERIC dropdown restoration for all sheets
            # self._extend_data_validations(ws, self.DATA_START_ROW_INDEX + 1)

        wb.save(self.output_path)
        print(f"✔️ GST Output written successfully → {self.output_path}")
