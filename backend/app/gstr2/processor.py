from openpyxl import load_workbook

from app.gstr1.utils.date_utils import parse_excel_or_date
from app.gstr1.utils.gst_utils import safe_string, to_float
from app.gstr2.reconciler import reconcile_b2b
from app.utils.logger import setup_logger

logger = setup_logger(
    name="gstr2.processor",
    log_file="logs/gstr2_processor.log"
)


# -------------------------------------------------
# Robust header detector
# -------------------------------------------------
def _detect_header_row(ws, required_headers, max_scan_rows=20):
    best_match = {"row": None, "score": 0, "col_map": {}}

    for row_idx in range(1, max_scan_rows):
        row = [safe_string(c.value) for c in ws[row_idx]]
        next_row = [safe_string(c.value) for c in ws[row_idx + 1]]

        col_map = {}
        score = 0

        for col_idx in range(len(row)):
            texts = []

            if col_idx < len(next_row) and next_row[col_idx]:
                texts.append(next_row[col_idx].lower())

            if row[col_idx]:
                texts.append(row[col_idx].lower())

            combined = " ".join(texts)

            for field, aliases in required_headers.items():
                for alias in aliases:
                    if alias.lower() in combined and field not in col_map:
                        col_map[field] = col_idx
                        score += 1

        if score > best_match["score"]:
            best_match = {
                "row": row_idx,
                "score": score,
                "col_map": col_map,
            }

    missing = set(required_headers) - set(best_match["col_map"])
    if missing:
        raise ValueError(
            f"Header detection failed. Missing: {missing}. "
            f"Detected: {best_match['col_map']}"
        )

    return best_match["row"] + 1, best_match["col_map"]


# -------------------------------------------------
# Purchase Register Reader (AGGREGATED)
# -------------------------------------------------
def _read_purchase_register(purchase_path: str) -> list[dict]:
    wb = load_workbook(purchase_path, data_only=True)
    ws = wb.active

    REQUIRED_HEADERS = {
        "gstin": {"GSTIN", "SUPPLIER GSTIN", "VENDOR GSTN"},
        "invoice_no": {"INVOICE NO", "BILL NUMBER", "DOCUMENT NO"},
        "invoice_date": {"INVOICE DATE", "DATE"},
        "taxable_value": {"TAXABLE VALUE", "TOTAL PRICE"},
    }

    header_row, col_map = _detect_header_row(ws, REQUIRED_HEADERS)
    logger.info(f"Purchase header detected at row {header_row}")

    # -------------------------------------------------
    # aggregation buckets
    # -------------------------------------------------
    aggregated_sum = {}     # key -> float
    aggregated_count = {}   # key -> int
    aggregated_lines = {}   # key -> list[float]

    for excel_row_no, r in enumerate(
        ws.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1,
    ):
        gstin = safe_string(r[col_map["gstin"]])
        invoice_no = safe_string(r[col_map["invoice_no"]])
        invoice_date = parse_excel_or_date(r[col_map["invoice_date"]])
        taxable_value = to_float(r[col_map["taxable_value"]]) or 0.0

        if not gstin or not invoice_no or not invoice_date:
            continue

        key = (gstin, invoice_no, invoice_date)

        if key not in aggregated_sum:
            aggregated_sum[key] = 0.0
            aggregated_count[key] = 0
            aggregated_lines[key] = []

        aggregated_sum[key] += taxable_value
        aggregated_count[key] += 1
        aggregated_lines[key].append(taxable_value)

        logger.debug(
            f"PURCHASE LINE | ROW={excel_row_no} | GSTIN={gstin} | "
            f"INV={invoice_no} | DATE={invoice_date} | TAXABLE={taxable_value}"
        )

    # -------------------------------------------------
    # convert aggregated → invoice-level rows
    # -------------------------------------------------
    rows = []

    for (gstin, invoice_no, invoice_date), total in aggregated_sum.items():
        count = aggregated_count[(gstin, invoice_no, invoice_date)]
        lines = aggregated_lines[(gstin, invoice_no, invoice_date)]

        logger.info(
            f"AGGREGATED PURCHASE INVOICE | GSTIN={gstin} | "
            f"INV={invoice_no} | DATE={invoice_date} | "
            f"ROWS={count} | "
            f"LINE_VALUES={lines} | "
            f"TOTAL_TAXABLE={round(total, 2)}"
        )

        rows.append({
            "gstin": gstin,
            "invoice_no": invoice_no,
            "invoice_date": invoice_date,
            "taxable_value": round(total, 2),
        })

    logger.info(
        f"Purchase aggregation completed | "
        f"RAW_ROWS={ws.max_row - header_row} | "
        f"INVOICE_ROWS={len(rows)}"
    )

    return rows



# -------------------------------------------------
# GSTR-2B B2B Reader (INVOICE LEVEL)
# -------------------------------------------------
def _read_gstr2b_b2b(gstr2b_path: str) -> list[dict]:
    wb = load_workbook(gstr2b_path, data_only=True)

    if "B2B" not in wb.sheetnames:
        raise ValueError("B2B sheet not found in GSTR-2B")

    ws = wb["B2B"]

    REQUIRED_HEADERS = {
        "gstin": {"gstin of supplier"},
        "invoice_no": {"invoice number"},
        "invoice_date": {"invoice date"},
        "taxable_value": {"taxable value"},
    }

    header_row, col_map = _detect_header_row(ws, REQUIRED_HEADERS)
    logger.info(f"GSTR2B header detected at row {header_row}")

    rows = []

    for r in ws.iter_rows(min_row=header_row + 1, values_only=True):
        gstin = safe_string(r[col_map["gstin"]])
        invoice_no = safe_string(r[col_map["invoice_no"]])

        if not gstin or not invoice_no:
            continue

        rows.append({
            "gstin": gstin,
            "invoice_no": invoice_no,
            "invoice_date": parse_excel_or_date(r[col_map["invoice_date"]]),
            "taxable_value": to_float(r[col_map["taxable_value"]]),
        })

    logger.info(f"GSTR2B rows read: {len(rows)}")
    return rows


# -------------------------------------------------
# Public Entry Point
# -------------------------------------------------
def process_b2b_single_state(purchase_path: str, gstr2b_path: str) -> dict:
    purchase_rows = _read_purchase_register(purchase_path)
    gstr2b_rows = _read_gstr2b_b2b(gstr2b_path)

    reconciled_rows = reconcile_b2b(
        gstr2b_rows=gstr2b_rows,
        purchase_rows=purchase_rows,
    )

    summary = {
        "total_rows": len(reconciled_rows),
        "matched": sum(r["comment"] == "Matched" for r in reconciled_rows),
        "not_matched": sum(r["comment"] == "Not Matched" for r in reconciled_rows),
        "not_in_books": sum(r["comment"] == "Not in Books" for r in reconciled_rows),
        "not_in_2b": sum(r["comment"] == "Not Found in 2B" for r in reconciled_rows),
    }

    return summary
