from openpyxl import load_workbook

from app.utils.logger import setup_logger
from app.gstr1.utils.gst_utils import safe_string, to_float
from app.gstr1.utils.date_utils import parse_excel_or_date
from app.gstr2.processor import detect_header_row
from app.gstr1.state_utils import state_code_from_value

logger = setup_logger(
    name="gstr2.b2b.reader",
    log_file="logs/gstr2_b2b_reader.log"
)

# -------------------------------------------------
# Purchase Register Reader (AGGREGATED + POS)
# -------------------------------------------------
def read_purchase_register(purchase_path: str) -> list[dict]:
    wb = load_workbook(purchase_path, data_only=True)
    ws = wb.active

    REQUIRED_HEADERS = {
        "gstin": {"GSTIN", "SUPPLIER GSTIN", "VENDOR GSTN"},
        "invoice_no": {"INVOICE NO", "BILL NUMBER", "DOCUMENT NO"},
        "invoice_date": {"INVOICE DATE", "DATE"},
        "taxable_value": {"TAXABLE VALUE", "TOTAL PRICE"},
        "place_of_supply": {"PLACE OF SUPPLY", "POS"},
    }

    header_row, col_map = detect_header_row(ws, REQUIRED_HEADERS)
    logger.info(f"Purchase header detected at row {header_row}")

    aggregated = {}

    for r in ws.iter_rows(min_row=header_row + 1, values_only=True):
        gstin = safe_string(r[col_map["gstin"]])
        invoice_no = safe_string(r[col_map["invoice_no"]])
        invoice_date = parse_excel_or_date(r[col_map["invoice_date"]])
        taxable_value = to_float(r[col_map["taxable_value"]]) or 0.0
        pos_state = state_code_from_value(r[col_map["place_of_supply"]])

        if not gstin or not invoice_no or not invoice_date or not pos_state:
            continue

        key = (gstin, invoice_no, invoice_date, pos_state)
        aggregated.setdefault(key, 0.0)
        aggregated[key] += taxable_value

    rows = []
    for (gstin, invoice_no, invoice_date, pos_state), total in aggregated.items():
        rows.append({
            "gstin": gstin,
            "invoice_no": invoice_no,
            "invoice_date": invoice_date,
            "taxable_value": round(total, 2),
            "pos_state": pos_state,
        })

    logger.info(f"Purchase invoices aggregated: {len(rows)}")
    return rows


# -------------------------------------------------
# GSTR-2B B2B Reader (INVOICE + POS)
# -------------------------------------------------
def read_gstr2b_b2b(gstr2b_path: str) -> list[dict]:
    wb = load_workbook(gstr2b_path, data_only=True)

    if "B2B" not in wb.sheetnames:
        raise ValueError("B2B sheet not found in GSTR-2B")

    ws = wb["B2B"]

    REQUIRED_HEADERS = {
        "gstin": {"gstin of supplier"},
        "invoice_no": {"invoice number"},
        "invoice_date": {"invoice date"},
        "taxable_value": {"taxable value"},
        "place_of_supply": {"place of supply"},
    }

    header_row, col_map = detect_header_row(ws, REQUIRED_HEADERS)
    logger.info(f"GSTR-2B header detected at row {header_row}")

    rows = []

    for r in ws.iter_rows(min_row=header_row + 1, values_only=True):
        gstin = safe_string(r[col_map["gstin"]])
        invoice_no = safe_string(r[col_map["invoice_no"]])
        pos_state = state_code_from_value(r[col_map["place_of_supply"]])

        if not gstin or not invoice_no or not pos_state:
            continue

        rows.append({
            "gstin": gstin,
            "invoice_no": invoice_no,
            "invoice_date": parse_excel_or_date(r[col_map["invoice_date"]]),
            "taxable_value": to_float(r[col_map["taxable_value"]]),
            "pos_state": pos_state,
        })

    logger.info(f"GSTR-2B rows read: {len(rows)}")
    return rows
