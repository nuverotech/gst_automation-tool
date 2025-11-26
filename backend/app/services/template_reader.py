import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import pandas as pd
from typing import Dict, List, Optional, Tuple
import re

from app.utils.logger import setup_logger

logger = setup_logger(__name__)


class TemplateReader:
    """
    Reads and analyzes an Excel template to understand its structure
    Then maps and processes user data to match that structure exactly
    """
    
    def __init__(self, template_path: str):
        """Initialize template reader"""
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"Template not found: {template_path}")
        
        self.template_path = template_path
        self.template_structure = {}
        self._analyze_template()
    
    def _analyze_template(self):
        """
        Thoroughly analyze the template structure
        Extract: sheets, headers, data ranges, merged cells, formatting, etc.
        """
        try:
            wb = load_workbook(self.template_path)
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                logger.info(f"Analyzing sheet: {sheet_name}")
                
                sheet_info = {
                    'name': sheet_name,
                    'headers': [],
                    'header_row': None,
                    'data_start_row': None,
                    'data_end_row': ws.max_row,
                    'columns': [],
                    'merged_cells': [],
                    'formatting': {},
                    'sample_row': None
                }
                
                # Find header row (first non-empty row or specific row)
                header_row = self._find_header_row(ws)
                if header_row:
                    sheet_info['header_row'] = header_row
                    sheet_info['headers'] = self._extract_headers(ws, header_row)
                    sheet_info['data_start_row'] = header_row + 1
                    
                    # Analyze columns
                    for col_idx, header in enumerate(sheet_info['headers'], 1):
                        col_info = {
                            'index': col_idx,
                            'letter': get_column_letter(col_idx),
                            'name': header,
                            'data_type': 'unknown',
                            'width': ws.column_dimensions[get_column_letter(col_idx)].width
                        }
                        sheet_info['columns'].append(col_info)
                    
                    # Extract sample row for type inference
                    if ws.max_row > header_row:
                        sheet_info['sample_row'] = self._extract_row(ws, header_row + 1, sheet_info['headers'])
                
                # Capture merged cells
                sheet_info['merged_cells'] = list(ws.merged_cells.ranges)
                
                # Capture formatting from header row
                if sheet_info['header_row']:
                    for col_idx, col_info in enumerate(sheet_info['columns'], 1):
                        cell = ws.cell(sheet_info['header_row'], col_idx)
                        sheet_info['formatting'][col_info['name']] = {
                            'font': cell.font.copy() if cell.font else None,
                            'fill': cell.fill.copy() if cell.fill else None,
                            'alignment': cell.alignment.copy() if cell.alignment else None,
                            'border': cell.border.copy() if cell.border else None,
                            'number_format': cell.number_format
                        }
                
                self.template_structure[sheet_name] = sheet_info
                logger.info(f"Sheet '{sheet_name}' analyzed. Headers: {sheet_info['headers']}")
            
            wb.close()
        
        except Exception as e:
            logger.error(f"Error analyzing template: {str(e)}", exc_info=True)
            raise
    
    def _find_header_row(self, ws) -> Optional[int]:
        """
        Find the header row in the worksheet
        Usually the first row with content, or explicitly marked row
        """
        for row_idx in range(1, min(ws.max_row + 1, 10)):  # Check first 10 rows
            row_values = [cell.value for cell in ws[row_idx]]
            if any(row_values):  # Row has content
                # Check if this looks like a header row (text values, not numbers)
                non_empty = [v for v in row_values if v is not None]
                if non_empty and all(isinstance(v, str) for v in non_empty):
                    return row_idx
        
        return 1  # Default to first row
    
    def _extract_headers(self, ws, header_row: int) -> List[str]:
        """Extract column headers from a specific row"""
        headers = []
        for cell in ws[header_row]:
            if cell.value:
                headers.append(str(cell.value).strip())
            else:
                # For empty cells, generate a generic name
                headers.append(f"Column_{len(headers) + 1}")
        
        return [h for h in headers if h and h != ""]
    
    def _extract_row(self, ws, row_idx: int, headers: List[str]) -> Dict[str, any]:
        """Extract a sample row data"""
        row_data = {}
        for col_idx, header in enumerate(headers, 1):
            try:
                cell = ws.cell(row_idx, col_idx)
                row_data[header] = cell.value
            except Exception as e:
                logger.warning(f"Error reading cell at row {row_idx}, col {col_idx}: {e}")
                row_data[header] = None
        return row_data
    
    def process_user_data(self, user_df: pd.DataFrame) -> Dict[str, pd.DataFrame]:
        """
        Process user's data and split it into sheets matching template structure
        
        Args:
            user_df: User's uploaded data (from CSV or Excel)
        
        Returns:
            Dict of {sheet_name: processed_dataframe}
        """
        processed_data = {}
        
        for sheet_name, sheet_info in self.template_structure.items():
            if not sheet_info['headers']:
                logger.warning(f"Sheet '{sheet_name}' has no headers, skipping")
                continue
            
            # Try to match user columns to template columns
            template_headers = sheet_info['headers']
            mapped_df = self._map_user_data_to_template(user_df, template_headers)
            
            if not mapped_df.empty:
                processed_data[sheet_name] = mapped_df
                logger.info(f"Mapped {len(mapped_df)} rows to sheet '{sheet_name}'")
            else:
                # Empty dataframe with template headers
                processed_data[sheet_name] = pd.DataFrame(columns=template_headers)
        
        return processed_data
    
    def _map_user_data_to_template(self, user_df: pd.DataFrame, template_headers: List[str]) -> pd.DataFrame:
        """
        Intelligently map user data columns to template columns
        """
        mapped_df = pd.DataFrame()
        
        for template_col in template_headers:
            # Strategy 1: Exact match
            if template_col in user_df.columns:
                mapped_df[template_col] = user_df[template_col]
                logger.info(f"[Exact] Mapped '{template_col}'")
                continue
            
            # Strategy 2: Case-insensitive match
            matching_cols = [col for col in user_df.columns 
                           if str(col).lower() == str(template_col).lower()]
            if matching_cols:
                mapped_df[template_col] = user_df[matching_cols[0]]
                logger.info(f"[Case-insensitive] Mapped '{matching_cols[0]}' -> '{template_col}'")
                continue
            
            # Strategy 3: Partial/fuzzy match
            matching_cols = [col for col in user_df.columns 
                           if self._fuzzy_match(str(template_col), str(col))]
            if matching_cols:
                mapped_df[template_col] = user_df[matching_cols[0]]
                logger.info(f"[Fuzzy] Mapped '{matching_cols[0]}' -> '{template_col}'")
                continue
            
            # Strategy 4: If template column contains keywords, try to infer
            inferred_col = self._infer_column(template_col, user_df.columns)
            if inferred_col:
                mapped_df[template_col] = user_df[inferred_col]
                logger.info(f"[Inferred] Mapped '{inferred_col}' -> '{template_col}'")
                continue
            
            # Not found - add empty column
            mapped_df[template_col] = None
            logger.warning(f"[Not Found] No match for template column '{template_col}'")
        
        return mapped_df
    
    def _fuzzy_match(self, str1: str, str2: str, threshold: float = 0.7) -> bool:
        """
        Simple fuzzy string matching
        """
        s1 = str1.lower().replace(' ', '').replace('_', '')
        s2 = str2.lower().replace(' ', '').replace('_', '')
        
        # Calculate similarity
        matching_chars = sum(1 for a, b in zip(s1, s2) if a == b)
        similarity = matching_chars / max(len(s1), len(s2))
        
        return similarity >= threshold
    
    def _infer_column(self, template_col: str, available_cols) -> Optional[str]:
        """
        Infer column based on keywords
        E.g., 'GST No' might match with 'GSTIN' or 'GST_Number'
        """
        keywords = {
            'gst': ['gstin', 'gst_no', 'gst number', 'gstnumber'],
            'invoice': ['invoice_no', 'inv_no', 'invoice number', 'bill_no'],
            'date': ['invoice_date', 'bill_date', 'date', 'trans_date'],
            'amount': ['value', 'amount', 'total', 'qty_value'],
            'quantity': ['qty', 'quantity', 'units'],
            'hsn': ['hsn_code', 'hsn', 'hsncode'],
            'sac': ['sac_code', 'sac', 'saccode'],
            'rate': ['tax_rate', 'rate', 'gst_rate'],
            'igst': ['igst', 'igst_value'],
            'sgst': ['sgst', 'sgst_value'],
            'cgst': ['cgst', 'cgst_value']
        }
        
        template_lower = template_col.lower()
        
        for keyword, patterns in keywords.items():
            if keyword in template_lower:
                for pattern in patterns:
                    for col in available_cols:
                        if pattern in col.lower().replace(' ', '').replace('_', ''):
                            return col
        
        return None
    
    def get_template_info(self) -> Dict:
        """
        Get readable information about the template
        """
        info = {}
        for sheet_name, sheet_info in self.template_structure.items():
            info[sheet_name] = {
                'headers': sheet_info['headers'],
                'columns_count': len(sheet_info['columns']),
                'max_rows': sheet_info['data_end_row']
            }
        return info
