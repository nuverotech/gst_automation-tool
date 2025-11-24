import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
import openpyxl.cell.cell
from typing import Dict, Optional
import pandas as pd

from app.config import settings
from app.utils.logger import setup_logger

logger = setup_logger(__name__)


class TemplateService:
    
    def __init__(self, custom_template_path: Optional[str] = None):
        """
        Initialize template service with optional custom template
        
        Args:
            custom_template_path: Path to user's custom template, or None for default
        """
        if custom_template_path and os.path.exists(custom_template_path):
            self.template_path = custom_template_path
            logger.info(f"Using custom template: {custom_template_path}")
        else:
            # Use default template
            self.template_path = os.path.join(
                settings.TEMPLATES_DIR,
                settings.DEFAULT_TEMPLATE_NAME
            )
            logger.info(f"Using default template: {self.template_path}")
    
    def load_template_structure(self):
        """
        Load the template and extract sheet names and column headers
        """
        try:
            wb = load_workbook(self.template_path)
            structure = {}
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Get headers from first row
                headers = []
                for cell in ws[1]:
                    if cell.value:
                        headers.append(cell.value)
                
                structure[sheet_name] = {
                    'headers': headers,
                    'row_count': ws.max_row
                }
                
                logger.info(f"Sheet '{sheet_name}' has columns: {headers}")
            
            wb.close()
            return structure
        
        except Exception as e:
            logger.error(f"Error loading template structure: {str(e)}", exc_info=True)
            raise
    
    def create_gst_file_from_template(self, output_path: str, data: Dict[str, pd.DataFrame]) -> str:
        """
        Create GST file by populating the template with processed data
        Follows the exact template structure, sheets, and column order
        
        Args:
            output_path: Path where the file will be saved
            data: Dict with sheet names as keys and DataFrames as values
        
        Returns:
            Path to the created file
        """
        try:
            # Load the template workbook
            wb = load_workbook(self.template_path)
            logger.info(f"Loaded template from: {self.template_path}")
            
            # Get template structure
            template_structure = self.load_template_structure()
            
            # Process each sheet in the template
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                logger.info(f"Processing sheet: {sheet_name}")
                
                # Get headers from template
                template_headers = []
                for cell in ws[1]:
                    if cell.value:
                        template_headers.append(str(cell.value).strip())
                
                logger.info(f"Template headers for '{sheet_name}': {template_headers}")
                
                # Check if we have data for this sheet
                if sheet_name in data and not data[sheet_name].empty:
                    df = data[sheet_name]
                    
                    logger.info(f"Found data for sheet '{sheet_name}': {len(df)} rows")
                    
                    # Map data columns to template columns
                    mapped_data = self._map_columns_to_template(df, template_headers, sheet_name)
                    
                    # Clear existing data rows (keep header) - handle merged cells
                    max_row = ws.max_row
                    if max_row > 1:  # Only if there are data rows beyond header
                        # Delete rows instead of clearing cells (avoids merged cell issues)
                        ws.delete_rows(2, max_row - 1)
                    
                    # Write mapped data to sheet
                    for row_idx, (_, row_data) in enumerate(mapped_data.iterrows(), start=2):
                        for col_idx, header in enumerate(template_headers, start=1):
                            cell = ws.cell(row=row_idx, column=col_idx)
                            value = row_data.get(header, None)
                            
                            # Only set value if it's not a merged cell
                            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                                cell.value = value
                    
                    logger.info(f"Wrote {len(mapped_data)} rows to sheet '{sheet_name}'")
                else:
                    # Clear data rows for empty sheets - handle merged cells
                    max_row = ws.max_row
                    if max_row > 1:  # Only if there are data rows beyond header
                        ws.delete_rows(2, max_row - 1)
                    
                    logger.info(f"Sheet '{sheet_name}' has no data, cleared existing rows")
            
            # Save the workbook
            wb.save(output_path)
            logger.info(f"GST template file saved to: {output_path}")
            wb.close()
            
            return output_path
        
        except Exception as e:
            logger.error(f"Error creating GST file from template: {str(e)}", exc_info=True)
            raise

    
    def _map_columns_to_template(self, df: pd.DataFrame, template_headers: list, sheet_name: str) -> pd.DataFrame:
        """
        Map DataFrame columns to template columns
        Try to match column names intelligently
        
        Args:
            df: Source DataFrame
            template_headers: List of template column headers
            sheet_name: Name of the sheet being processed
        
        Returns:
            DataFrame with columns in template order
        """
        mapped_df = pd.DataFrame()
        
        for template_col in template_headers:
            # Try exact match
            if template_col in df.columns:
                mapped_df[template_col] = df[template_col]
            else:
                # Try case-insensitive match
                matching_cols = [col for col in df.columns if str(col).lower() == str(template_col).lower()]
                if matching_cols:
                    mapped_df[template_col] = df[matching_cols[0]]
                    logger.info(f"Mapped '{matching_cols[0]}' -> '{template_col}'")
                else:
                    # Try partial match
                    matching_cols = [col for col in df.columns if str(template_col).lower() in str(col).lower()]
                    if matching_cols:
                        mapped_df[template_col] = df[matching_cols[0]]
                        logger.info(f"Partial mapped '{matching_cols[0]}' -> '{template_col}'")
                    else:
                        # Column not found, add empty column
                        mapped_df[template_col] = None
                        logger.warning(f"No match found for template column '{template_col}' in sheet '{sheet_name}'")
        
        return mapped_df
    
    def get_template_sheets(self) -> list:
        """
        Get all sheet names from the template
        """
        try:
            wb = load_workbook(self.template_path)
            sheets = wb.sheetnames
            wb.close()
            return sheets
        except Exception as e:
            logger.error(f"Error getting template sheets: {str(e)}")
            return []
    
    @staticmethod
    def save_user_template(file_content: bytes, user_id: int, filename: str) -> str:
        """
        Save user's custom template
        
        Args:
            file_content: Template file content
            user_id: User ID
            filename: Original filename
        
        Returns:
            Path to saved template
        """
        try:
            # Create user templates directory if not exists
            user_templates_dir = settings.USER_TEMPLATES_DIR
            os.makedirs(user_templates_dir, exist_ok=True)
            
            # Generate unique filename
            from app.utils.helpers import generate_unique_filename
            unique_filename = f"user_{user_id}_{generate_unique_filename(filename)}"
            template_path = os.path.join(user_templates_dir, unique_filename)
            
            # Save file
            with open(template_path, 'wb') as f:
                f.write(file_content)
            
            logger.info(f"Saved custom template for user {user_id}: {template_path}")
            return template_path
        
        except Exception as e:
            logger.error(f"Error saving user template: {str(e)}", exc_info=True)
            raise
